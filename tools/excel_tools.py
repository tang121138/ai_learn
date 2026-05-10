"""Excel 读写工具 — 读取和写入 Excel 文件"""
import os
import json
from backend.logger import get_logger

logger = get_logger("tool.excel")

try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False
    logger.warning("openpyxl 未安装，Excel 工具不可用")

SAFE_BASE_DIR = os.path.expanduser("~")


def _safe_path(path: str) -> str:
    real_path = os.path.realpath(os.path.expanduser(path))
    if not real_path.startswith(SAFE_BASE_DIR):
        raise PermissionError(f"禁止访问用户目录之外的路径: {path}")
    return real_path


def read_excel(file_path: str, sheet: str = "", max_rows: int = 100) -> str:
    """读取 Excel 文件内容。

    Args:
        file_path: Excel 文件路径 (.xlsx/.xls)
        sheet: 工作表名称，默认为第一个工作表
        max_rows: 最大读取行数，默认 100
    """
    if not HAS_OPENPYXL:
        return "错误: openpyxl 未安装，请运行: pip install openpyxl"

    try:
        path = _safe_path(file_path)
    except PermissionError as e:
        return str(e)

    if not os.path.exists(path):
        return f"错误: 文件不存在: {path}"

    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb[sheet] if sheet else wb.active
        if ws is None:
            wb.close()
            return "错误: 工作簿中没有工作表"

        rows = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i >= max_rows:
                rows.append(("...",) * len(row))
                break
            rows.append(row)

        wb.close()

        if not rows:
            return "Excel 文件为空。"

        result = []
        result.append(f"工作表: {ws.title}, 共读取 {len(rows)} 行")
        for i, row in enumerate(rows, 1):
            result.append(f"  [{i}] {' | '.join(str(c) if c is not None else '' for c in row)}")

        return "\n".join(result)
    except Exception as e:
        return f"读取 Excel 失败: {e}"


def write_excel(data: str, output_file: str) -> str:
    """将 JSON 数据写入 Excel 文件。

    Args:
        data: JSON 格式的二维数组，例如: [["姓名","年龄"],["张三",25],["李四",30]]
        output_file: 输出文件名（保存在当前目录）
    """
    if not HAS_OPENPYXL:
        return "错误: openpyxl 未安装，请运行: pip install openpyxl"

    try:
        rows = json.loads(data)
        if not isinstance(rows, list) or not all(isinstance(r, list) for r in rows):
            return "错误: 数据格式不正确，应为 JSON 二维数组"
    except json.JSONDecodeError as e:
        return f"错误: 数据 JSON 解析失败: {e}"

    try:
        path = os.path.join(os.path.expanduser("~"), output_file)
    except Exception:
        path = output_file

    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        for row in rows:
            ws.append(row)
        wb.save(path)
        return f"Excel 文件已保存: {path} (共 {len(rows)} 行)"
    except Exception as e:
        return f"写入 Excel 失败: {e}"


tool_defs = [
    {
        "type": "function",
        "function": {
            "name": "read_excel",
            "description": "读取 Excel 文件 (.xlsx/.xls) 的内容。用于查看电子表格数据。",
            "parameters": {
                "type": "object",
                "properties": {
                    "file_path": {
                        "type": "string",
                        "description": "Excel 文件路径",
                    },
                    "sheet": {
                        "type": "string",
                        "description": "工作表名称，不指定则读取第一个",
                    },
                    "max_rows": {
                        "type": "integer",
                        "description": "最大读取行数，默认 100",
                    },
                },
                "required": ["file_path"],
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "write_excel",
            "description": "将数据写入 Excel 文件。用于导出表格数据、生成报表。",
            "parameters": {
                "type": "object",
                "properties": {
                    "data": {
                        "type": "string",
                        "description": "JSON 格式的二维数组，例如: [[\"姓名\",\"年龄\"],[\"张三\",25],[\"李四\",30]]",
                    },
                    "output_file": {
                        "type": "string",
                        "description": "输出文件名，例如: report.xlsx。保存在用户主目录。",
                    },
                },
                "required": ["data", "output_file"],
            },
        },
    },
]
